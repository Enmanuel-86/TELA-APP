import calendar
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Tuple, Dict, Optional, Any
from configuraciones.configuracion import app_configuracion
from reportes.reporte_base import ReporteBase
from excepciones.base_datos_error import BaseDatosError
from recursos_graficos_y_logicos.utilidades.funciones_sistema import FuncionSistema
from recursos_graficos_y_logicos.utilidades.base_de_datos import asistencia_empleado_servicio


class ReporteAsistenciaMensualEmpleados(ReporteBase):
    def __init__(self):
        self.RUTA_REPORTES_ASISTENCIA = app_configuracion.DIRECTORIO_REPORTES_EMPLEADOS
    
    def transformar_lista_asistencia_inasistencia_empleados(self, lista_asistencia_inasistencia_empleados: List, lista_dia_mes_con_semana: List, anio: int, mes: int)-> Optional[List[Dict]]:
        lista_dict_asistencia_inasistencia_empleados = []
    
        for dia_numero_str, dia_semana in lista_dia_mes_con_semana:
            dia_numero = int(dia_numero_str)
            fecha_actual = datetime.datetime(anio, mes, dia_numero).date()
            
            registro_encontrado = None
            for registro_bd in lista_asistencia_inasistencia_empleados:
                if registro_bd[1] == str(fecha_actual):
                    registro_encontrado = registro_bd
                    break
            
            fila_tabla = {
                "Dia_Numero": dia_numero_str,
                "Dia_Semana": dia_semana,
                "Es_Feriado": None,
                "Hombres_Presentes": 0,
                "Mujeres_Presentes": 0,
                "Total_Presentes": 0,
                "Hombres_Ausentes": 0,
                "Mujeres_Ausentes": 0,
                "Total_Ausentes": 0
            }
            
            if registro_encontrado:
                fila_tabla["Hombres_Presentes"] = registro_encontrado[2]
                fila_tabla["Mujeres_Presentes"] = registro_encontrado[3]
                fila_tabla["Total_Presentes"] = registro_encontrado[4]
                fila_tabla["Hombres_Ausentes"] = registro_encontrado[5]
                fila_tabla["Mujeres_Ausentes"] = registro_encontrado[6]
                fila_tabla["Total_Ausentes"] = registro_encontrado[7]
            else:
                if dia_semana in ["Sábado", "Domingo"]:
                    fila_tabla["Es_Feriado"] = "Fin de Semana"
            
            lista_dict_asistencia_inasistencia_empleados.append(fila_tabla)
        return lista_dict_asistencia_inasistencia_empleados
    
    def transformar_lista_totales_asistencia_inasistencia_empleados(self, tupla_asistencia_inasistencia_empleados: Tuple) -> Optional[List[Dict]]:
        lista_dict_asistencia_inasistencia_empleados = []
    
        sumatorio_hombres_presentes = tupla_asistencia_inasistencia_empleados[1]
        sumatoria_mujeres_presentes = tupla_asistencia_inasistencia_empleados[2]
        sumatoria_general_presentes = tupla_asistencia_inasistencia_empleados[3]
            
        sumatoria_hombres_ausentes = tupla_asistencia_inasistencia_empleados[4]
        sumatoria_mujeres_ausentes = tupla_asistencia_inasistencia_empleados[5]
        sumatoria_general_ausentes = tupla_asistencia_inasistencia_empleados[6]
            
        fila_tabla = {
            "Sumatoria_Hombres_Presentes": sumatorio_hombres_presentes,
            "Sumatoria_Mujeres_Presentes": sumatoria_mujeres_presentes,
            "Sumatoria_General_Presentes": sumatoria_general_presentes,
            "Sumatoria_Hombres_Ausentes": sumatoria_hombres_ausentes,
            "Sumatoria_Mujeres_Ausentes": sumatoria_mujeres_ausentes,
            "Sumatoria_General_Ausentes": sumatoria_general_ausentes
        }
            
        lista_dict_asistencia_inasistencia_empleados.append(fila_tabla)
        
        return lista_dict_asistencia_inasistencia_empleados
    
    def transformar_promedio_asistencia_inasistencia_empleados(self, tupla_promedio_asistencia_inasistencia_empleados: Tuple) -> Optional[List[Dict]]:
        lista_dict_promedio_asistencia_inasistencia_empleados = []
    
        promedio_hombres_presentes = tupla_promedio_asistencia_inasistencia_empleados[1]
        promedio_mujeres_presentes = tupla_promedio_asistencia_inasistencia_empleados[2]
        promedio_general_presentes = tupla_promedio_asistencia_inasistencia_empleados[3]
        
        promedio_hombres_ausentes = tupla_promedio_asistencia_inasistencia_empleados[4]
        promedio_mujeres_ausentes = tupla_promedio_asistencia_inasistencia_empleados[5]
        promedio_general_ausentes = tupla_promedio_asistencia_inasistencia_empleados[6]
        
        fila_tabla = {
            "Promedio_Hombres_Presentes": promedio_hombres_presentes,
            "Promedio_Mujeres_Presentes": promedio_mujeres_presentes,
            "Promedio_General_Presentes": promedio_general_presentes,
            "Promedio_Hombres_Ausentes": promedio_hombres_ausentes,
            "Promedio_Mujeres_Ausentes": promedio_mujeres_ausentes,
            "Promedio_General_Ausentes": promedio_general_ausentes
        }
        
        lista_dict_promedio_asistencia_inasistencia_empleados.append(fila_tabla)
        
        return lista_dict_promedio_asistencia_inasistencia_empleados
    
    def transformar_porcentaje_asistencia_inasistencia_empleados(self, tupla_porcentaje_asistencia_inasistencia_empleados: Tuple) -> Optional[List[Dict]]:
        lista_dict_porcentaje_asistencia_inasistencia_empleados = []
    
        porcentaje_hombres_presentes = tupla_porcentaje_asistencia_inasistencia_empleados[1]
        porcentaje_mujeres_presentes = tupla_porcentaje_asistencia_inasistencia_empleados[2]
        porcentaje_general_presentes = tupla_porcentaje_asistencia_inasistencia_empleados[3]
        
        porcentaje_hombres_ausentes = tupla_porcentaje_asistencia_inasistencia_empleados[4]
        porcentaje_mujeres_ausentes = tupla_porcentaje_asistencia_inasistencia_empleados[5]
        porcentaje_general_ausentes = tupla_porcentaje_asistencia_inasistencia_empleados[6]
        
        fila_tabla = {
            "Porcentaje_Hombres_Presentes": porcentaje_hombres_presentes,
            "Porcentaje_Mujeres_Presentes": porcentaje_mujeres_presentes,
            "Porcentaje_General_Presentes": porcentaje_general_presentes,
            "Porcentaje_Hombres_Ausentes": porcentaje_hombres_ausentes,
            "Porcentaje_Mujeres_Ausentes": porcentaje_mujeres_ausentes,
            "Porcentaje_General_Ausentes": porcentaje_general_ausentes
        }
        
        lista_dict_porcentaje_asistencia_inasistencia_empleados.append(fila_tabla)
        
        return lista_dict_porcentaje_asistencia_inasistencia_empleados
    
    def cargar_dia_mes_con_dia_semana(self, anio: int, mes: int) -> List[Tuple]:
        num_dias = calendar.monthrange(anio, mes)[1]
        dias_con_semana = []
        diccionario_dias_semana_espaniol = {
            "Monday": "Lunes",
            "Tuesday": "Martes",
            "Wednesday": "Miercoles",
            "Thursday": "Jueves",
            "Friday": "Viernes",
            "Saturday": "Sábado",
            "Sunday": "Domingo"
        }
        
        for dia in range(1, num_dias + 1):
            fecha = datetime.date(anio, mes, dia)
            fecha_formateada = fecha.strftime("%d")
            
            dia_semana_ingles = calendar.day_name[fecha.weekday()]
            dia_semana_espaniol = diccionario_dias_semana_espaniol[dia_semana_ingles]
            
            dias_con_semana.append((fecha_formateada, dia_semana_espaniol))
        return dias_con_semana
    
    def cargar_meses_y_anio(self, anio: int) -> List[str]:
        meses_espaniol_y_anio = []
        meses_ingles = list(calendar.month_name)
        
        diccionario_meses_espaniol = {
            "January": "Enero",
            "February": "Febrero",
            "March": "Marzo",
            "April": "Abril",
            "May": "Mayo",
            "June": "Junio",
            "July": "Julio",
            "August": "Agosto",
            "September": "Septiembre",
            "October": "Octubre",
            "November": "Noviembre",
            "December": "Diciembre"
        }
        
        for mes in range(1, 13):
            mes_ingles = meses_ingles[mes]
            mes_espaniol = diccionario_meses_espaniol[mes_ingles]
            meses_espaniol_y_anio.append(mes_espaniol)
        meses_espaniol_y_anio.append(anio)
        
        return meses_espaniol_y_anio
    
    def cargar_datos(
        self,
        anio: int,
        mes: int
    ) -> Optional[List[Any]]:
        try:
            datos = []
            
            if (mes > 9):
                anio_mes = f"{anio}-{mes}"
            else:
                anio_mes = f"{anio}-0{mes}"
            
            # Posición 0: Día del mes con el día de semana
            lista_dia_mes_con_dia_semana = self.cargar_dia_mes_con_dia_semana(anio, mes)
            datos.append(lista_dia_mes_con_dia_semana)
            
            # Posición 1: Mes específico
            mes_especifico = self.cargar_meses_y_anio(anio)[(mes - 1)]
            datos.append(mes_especifico)
            
            # Posición 2: Año específico
            anio_especifico = self.cargar_meses_y_anio(anio)[12]
            datos.append(anio_especifico)
            
            if not(asistencia_empleado_servicio.obtener_conteo_asistencia_mensual(anio_mes)):
                raise BaseDatosError("ASISTENCIA_EMPLEADOS_NO_EXISTE", "No existen registros de asistencia de empleados en este año y mes.")
            
            # Posición 3: Lista Dict de la asistencia mensual de empleados
            lista_asistencia_inasistencia_empleados = asistencia_empleado_servicio.obtener_conteo_asistencia_mensual(anio_mes)
            lista_dict_asistencia_inasistencia_empleados = self.transformar_lista_asistencia_inasistencia_empleados(
                lista_asistencia_inasistencia_empleados,
                lista_dia_mes_con_dia_semana,
                anio,
                mes
            )
            datos.append(lista_dict_asistencia_inasistencia_empleados)
            
            # Posición 4: Lista Dict de la sumatoria de asistencias e inasistencias de empleados mensual
            tupla_sumatoria_asistencia_inasistencia_empleados = asistencia_empleado_servicio.obtener_sumatoria_asistencia_inasistencia(anio_mes)
            lista_dict_sumatoria_asistencia_inasistencia_empleados = self.transformar_lista_totales_asistencia_inasistencia_empleados(
                tupla_sumatoria_asistencia_inasistencia_empleados
            )
            datos.append(lista_dict_sumatoria_asistencia_inasistencia_empleados)
            
            # Posición 5: Lista Dict del promedio de asistencias e inasistencias de empleados mensual
            tupla_promedio_asistencia_inasistencia_empleados = asistencia_empleado_servicio.obtener_promedio_asistencia_inasistencia(anio_mes)
            lista_dict_promedio_asistencia_inasistencia_empleados = self.transformar_promedio_asistencia_inasistencia_empleados(
                tupla_promedio_asistencia_inasistencia_empleados
            )
            datos.append(lista_dict_promedio_asistencia_inasistencia_empleados)
            
            # Posición 6: Lista Dict del porcentaje de asistencias e inasistencias de empleados mensual
            tupla_porcentaje_asistencia_inasistencia_empleados = asistencia_empleado_servicio.obtener_porcentaje_asistencia_inasistenica(anio_mes)
            lista_dict_porcentaje_asistencia_inasistencia_empleados = self.transformar_porcentaje_asistencia_inasistencia_empleados(
                tupla_porcentaje_asistencia_inasistencia_empleados
            )
            datos.append(lista_dict_porcentaje_asistencia_inasistencia_empleados)
            
            # Posición 7: Total de Días hábiles
            total_dias_hablie = asistencia_empleado_servicio.obtener_dias_habiles(anio_mes)
            datos.append(total_dias_hablie)
            
            return datos
        except BaseDatosError as error:
            raise error
    
    def crear_libro_y_hoja(self):
        libro = Workbook()
        hoja = libro.active
        
        return libro, hoja
    
    def cargar_configuraciones_excel(self):
        fuente_negrita = Font(bold = True)
        alineacion_centrada = Alignment(horizontal = "center", vertical = "center")
        fuente_titulo = Font(size = 20, bold = True)
        relleno_encabezados = PatternFill(start_color = "92D050", end_color = "92D050", fill_type = "solid")
        relleno_fines_semana = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")

        tipo_borde = Side(border_style="thin", color="000000")
        borde_celda = Border(
            left = tipo_borde,
            right = tipo_borde,
            top = tipo_borde,
            bottom = tipo_borde
        )
        
        return fuente_negrita, alineacion_centrada, fuente_titulo, relleno_encabezados, relleno_fines_semana, borde_celda
    
    def cargar_encabezados_asistencia_mensual(self, hoja, borde_celda, relleno_encabezados, fuente_negrita, fuente_titulo, alineacion_centrada):
        hoja["B1"] = "ASISTENCIA GENERAL DE EMPLEADOS"
        hoja["A3"] = "Fecha"
        hoja["B3"] = "Día"
        hoja["C3"] = "Asistencia"
        hoja["F3"] = "Inasistencia"

        hoja["B1"].border = borde_celda
        hoja["A3"].border = borde_celda
        hoja["B3"].border = borde_celda
        hoja["C3"].border = borde_celda
        hoja["F3"].border = borde_celda

        hoja["B1"].fill = relleno_encabezados
        hoja["A3"].fill = relleno_encabezados
        hoja["B3"].fill = relleno_encabezados
        hoja["C3"].fill = relleno_encabezados
        hoja["F3"].fill = relleno_encabezados

        hoja.merge_cells("B1:G1")
        hoja.merge_cells("A3:A4")
        hoja.merge_cells("B3:B4")
        hoja.merge_cells("C3:E3")
        hoja.merge_cells("F3:H3")

        hoja["C4"] = "V"
        hoja["D4"] = "H"
        hoja["E4"] = "T"

        hoja["C4"].border = borde_celda
        hoja["D4"].border = borde_celda
        hoja["E4"].border = borde_celda

        hoja["F4"] = "V"
        hoja["G4"] = "H"
        hoja["H4"] = "T"

        hoja["F4"].border = borde_celda
        hoja["G4"].border = borde_celda
        hoja["H4"].border = borde_celda


        for num_fila in range(1, 5):
            for num_columna in range(1, 9):
                celda = hoja.cell(row = num_fila, column = num_columna)
                
                celda.font = fuente_negrita
                celda.alignment = alineacion_centrada
                
        hoja["B1"].font = fuente_titulo
        hoja["B1"].border = borde_celda

        for columna in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            hoja.column_dimensions[columna].width = 12
    
    def cargar_datos_asistencia_inasistencia_empleados(self, hoja, lista_dict_asistencia_mensual_empleados: List[Dict], borde_celda, alineacion_centrada, relleno_fines_semana):
        fila_actual = 5

        for fila_data in lista_dict_asistencia_mensual_empleados:
            # Escribimos el número de día
            hoja.cell(row = fila_actual, column = 1, value = fila_data["Dia_Numero"])
            hoja.cell(row = fila_actual, column = 1).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 1).border = borde_celda
            
            # Escribimos el día de semana
            hoja.cell(row = fila_actual, column = 2, value = fila_data["Dia_Semana"])
            hoja.cell(row = fila_actual, column = 2).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 2).border = borde_celda
            
            
            # Manejo de fin de semana
            if fila_data["Es_Feriado"] == "Fin de Semana":
                rango = f"C{fila_actual}:H{fila_actual}"
                hoja.merge_cells(start_row = fila_actual, start_column = 3, end_row = fila_actual, end_column = 8)
                hoja.cell(row = fila_actual, column = 3, value = "").fill = relleno_fines_semana
                hoja.cell(row = fila_actual, column = 3).border = borde_celda
                
                hoja.cell(row = fila_actual, column = 3).alignment = alineacion_centrada
                FuncionSistema.aplicar_borde_a_rango(hoja, rango, borde_celda)
            else:
                hoja.cell(row = fila_actual, column = 3, value = fila_data["Hombres_Presentes"])
                hoja.cell(row = fila_actual, column = 3).alignment = alineacion_centrada
                hoja.cell(row = fila_actual, column = 3).border = borde_celda
                
                hoja.cell(row = fila_actual, column = 4, value = fila_data["Mujeres_Presentes"])
                hoja.cell(row = fila_actual, column = 4).alignment = alineacion_centrada
                hoja.cell(row = fila_actual, column = 4).border = borde_celda
                
                hoja.cell(row = fila_actual, column = 5, value = fila_data["Total_Presentes"])
                hoja.cell(row = fila_actual, column = 5).alignment = alineacion_centrada
                hoja.cell(row = fila_actual, column = 5).border = borde_celda
                
                hoja.cell(row = fila_actual, column = 6, value = fila_data["Hombres_Ausentes"])
                hoja.cell(row = fila_actual, column = 6).alignment = alineacion_centrada
                hoja.cell(row = fila_actual, column = 6).border = borde_celda
                
                hoja.cell(row = fila_actual, column = 7, value = fila_data["Mujeres_Ausentes"])
                hoja.cell(row = fila_actual, column = 7).alignment = alineacion_centrada
                hoja.cell(row = fila_actual, column = 7).border = borde_celda
                
                hoja.cell(row = fila_actual, column = 8, value = fila_data["Total_Ausentes"])
                hoja.cell(row = fila_actual, column = 8).alignment = alineacion_centrada
                hoja.cell(row = fila_actual, column = 8).border = borde_celda
            
            fila_actual += 1
    
    def cargar_datos_sumatoria_asistencia_inasistencia_empleados(self, hoja, lista_dict_sumatoria_asistencia_inasistencia_empleados: List[Dict], alineacion_centrada, borde_celda):
        ultima_fila = hoja.max_row
        siguiente_fila_disponible = ultima_fila + 1

        hoja[f"A{siguiente_fila_disponible}"] = "Total"

        hoja[f"A{siguiente_fila_disponible}"].alignment = alineacion_centrada
        hoja[f"A{siguiente_fila_disponible}"].border = borde_celda
        hoja.row_dimensions[siguiente_fila_disponible].height = 25


        for fila_data in lista_dict_sumatoria_asistencia_inasistencia_empleados:
            hoja.cell(row = siguiente_fila_disponible, column = 3, value = fila_data["Sumatoria_Hombres_Presentes"])
            hoja.cell(row = siguiente_fila_disponible, column = 3).alignment = alineacion_centrada
            hoja.cell(row = siguiente_fila_disponible, column = 3).border = borde_celda
            
            hoja.cell(row = siguiente_fila_disponible, column = 4, value = fila_data["Sumatoria_Mujeres_Presentes"])
            hoja.cell(row = siguiente_fila_disponible, column = 4).alignment = alineacion_centrada
            hoja.cell(row = siguiente_fila_disponible, column = 4).border = borde_celda
            
            hoja.cell(row = siguiente_fila_disponible, column = 5, value = fila_data["Sumatoria_General_Presentes"])
            hoja.cell(row = siguiente_fila_disponible, column = 5).alignment = alineacion_centrada
            hoja.cell(row = siguiente_fila_disponible, column = 5).border = borde_celda
            
            hoja.cell(row = siguiente_fila_disponible, column = 6, value = fila_data["Sumatoria_Hombres_Ausentes"])
            hoja.cell(row = siguiente_fila_disponible, column = 6).alignment = alineacion_centrada
            hoja.cell(row = siguiente_fila_disponible, column = 6).border = borde_celda
            
            hoja.cell(row = siguiente_fila_disponible, column = 7, value = fila_data["Sumatoria_Mujeres_Ausentes"])
            hoja.cell(row = siguiente_fila_disponible, column = 7).alignment = alineacion_centrada
            hoja.cell(row = siguiente_fila_disponible, column = 7).border = borde_celda
            
            hoja.cell(row = siguiente_fila_disponible, column = 8, value = fila_data["Sumatoria_General_Ausentes"])
            hoja.cell(row = siguiente_fila_disponible, column = 8).alignment = alineacion_centrada
            hoja.cell(row = siguiente_fila_disponible, column = 8).border = borde_celda


        hoja.merge_cells(f"A{siguiente_fila_disponible}:B{siguiente_fila_disponible}")
        
        rango = f"A{siguiente_fila_disponible}:B{siguiente_fila_disponible}"
        FuncionSistema.aplicar_borde_a_rango(hoja, rango, borde_celda)
    
    def cargar_datos_dias_habiles(self, hoja, dias_habiles: int, fuente_negrita, alineacion_centrada, borde_celda, relleno_encabezados):
        hoja["J22"] = "Días Hábiles"
        hoja["J22"].font = fuente_negrita
        hoja["J22"].alignment = alineacion_centrada
        hoja["J22"].border = borde_celda
        hoja["J22"].fill = relleno_encabezados

        hoja.merge_cells("J22:K22")

        hoja["L22"] = dias_habiles
        hoja["L22"].alignment = alineacion_centrada
        hoja["L22"].border = borde_celda
    
    def cargar_encabezados_promedio_porcentaje_asistencia_inasistencia_empleados(self, hoja, fuente_negrita, alineacion_centrada, borde_celda, relleno_encabezados):
        hoja["L27"] = "Asistencia"
        hoja["L27"].font = fuente_negrita
        hoja["L27"].alignment = alineacion_centrada
        hoja["L27"].border = borde_celda
        hoja["L27"].fill = relleno_encabezados

        hoja["L28"] = "V"
        hoja["M28"] = "H"
        hoja["N28"] = "T"
        hoja["L28"].font = fuente_negrita
        hoja["M28"].font = fuente_negrita
        hoja["N28"].font = fuente_negrita
        hoja["L28"].alignment = alineacion_centrada
        hoja["M28"].alignment = alineacion_centrada
        hoja["N28"].alignment = alineacion_centrada
        hoja["L28"].border = borde_celda
        hoja["M28"].border = borde_celda
        hoja["N28"].border = borde_celda


        hoja["P27"] = "Inasistencia"
        hoja["P27"].font = fuente_negrita
        hoja["P27"].alignment = alineacion_centrada
        hoja["P27"].border = borde_celda
        hoja["P27"].fill = relleno_encabezados

        hoja["P28"] = "V"
        hoja["Q28"] = "H"
        hoja["R28"] = "T"
        hoja["P28"].font = fuente_negrita
        hoja["Q28"].font = fuente_negrita
        hoja["R28"].font = fuente_negrita
        hoja["P28"].alignment = alineacion_centrada
        hoja["Q28"].alignment = alineacion_centrada
        hoja["R28"].alignment = alineacion_centrada
        hoja["P28"].border = borde_celda
        hoja["Q28"].border = borde_celda
        hoja["R28"].border = borde_celda


        hoja.merge_cells("L27:N27")
        hoja.merge_cells("P27:R27")


        hoja["J28"] = "Indicadores"
        hoja["J28"].font = fuente_negrita
        hoja["J28"].border = borde_celda
        hoja["J28"].fill = relleno_encabezados


        hoja["J29"] = "Promedio"
        hoja["J29"].font = fuente_negrita
        hoja["J29"].border = borde_celda
        hoja["J29"].fill = relleno_encabezados

        hoja["J30"] = "Porcentaje"
        hoja["J30"].font = fuente_negrita
        hoja["J30"].border = borde_celda
        hoja["J30"].fill = relleno_encabezados


        hoja.merge_cells("J28:K28")
        hoja.merge_cells("J29:K29")
        hoja.merge_cells("J30:K30")
        
        rangos = ["L27:N27", "P27:R27","J28:K28", "J29:K29", "J30:K30"]
        
        for rango in rangos:
            FuncionSistema.aplicar_borde_a_rango(hoja, rango, borde_celda)
    
    def cargar_datos_promedio_asistencia_inasistencia_empleados(self, hoja, lista_dict_promedio_asistencia_inasistencia_empleados: List[Dict], alineacion_centrada, borde_celda):
        fila_actual = 29

        for fila_data in lista_dict_promedio_asistencia_inasistencia_empleados:
            hoja.cell(row = fila_actual, column = 12, value = fila_data["Promedio_Hombres_Presentes"])
            hoja.cell(row = fila_actual, column = 12).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 12).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 13, value = fila_data["Promedio_Mujeres_Presentes"])
            hoja.cell(row = fila_actual, column = 13).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 13).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 14, value = fila_data["Promedio_General_Presentes"])
            hoja.cell(row = fila_actual, column = 14).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 14).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 16, value = fila_data["Promedio_Hombres_Ausentes"])
            hoja.cell(row = fila_actual, column = 16).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 16).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 17, value = fila_data["Promedio_Mujeres_Ausentes"])
            hoja.cell(row = fila_actual, column = 17).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 17).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 18, value = fila_data["Promedio_General_Ausentes"])
            hoja.cell(row = fila_actual, column = 18).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 18).border = borde_celda
            
            fila_actual += 1
    
    def cargar_datos_porcentaje_asistencia_inasistencia_empleados(self, hoja, lista_dict_porcentaje_asistencia_inasistencia_empleados: List[Dict], alineacion_centrada, borde_celda):
        fila_actual = 30

        for fila_data in lista_dict_porcentaje_asistencia_inasistencia_empleados:
            hoja.cell(row = fila_actual, column = 12, value = fila_data["Porcentaje_Hombres_Presentes"])
            hoja.cell(row = fila_actual, column = 12).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 12).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 13, value = fila_data["Porcentaje_Mujeres_Presentes"])
            hoja.cell(row = fila_actual, column = 13).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 13).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 14, value = fila_data["Porcentaje_General_Presentes"])
            hoja.cell(row = fila_actual, column = 14).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 14).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 16, value = fila_data["Porcentaje_Hombres_Ausentes"])
            hoja.cell(row = fila_actual, column = 16).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 16).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 17, value = fila_data["Porcentaje_Mujeres_Ausentes"])
            hoja.cell(row = fila_actual, column = 17).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 17).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 18, value = fila_data["Porcentaje_General_Ausentes"])
            hoja.cell(row = fila_actual, column = 18).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 18).border = borde_celda
            
            fila_actual += 1
    
    def exportar(self, datos):
        try:
            self.RUTA_REPORTES_ASISTENCIA.mkdir(exist_ok = True)
                
            fuente_negrita = self.cargar_configuraciones_excel()[0]
            alineacion_centrada = self.cargar_configuraciones_excel()[1]
            fuente_titulo = self.cargar_configuraciones_excel()[2]
            relleno_encabezados = self.cargar_configuraciones_excel()[3]
            relleno_fines_semana = self.cargar_configuraciones_excel()[4]
            borde_celda = self.cargar_configuraciones_excel()[5]
                
            mes_especifico = datos[1]
            anio_especifico = datos[2]
            lista_dict_asistencia_mensual_empleados = datos[3]
            lista_dict_sumatoria_asistencia_inasistencia_empleados = datos[4]
            lista_dict_promedio_asistencia_inasistencia_empleados = datos[5]
            lista_dict_porcentaje_asistencia_inasistencia_empleados = datos[6]
            dias_habiles = datos[7]
            
            libro, hoja = self.crear_libro_y_hoja()
            
            hoja.title = f"ASISTENCIA {mes_especifico.upper()}-{anio_especifico}"
            nombre_archivo = f"REPORTE_ASISTENCIA_GENERAL_EMPLEADOS_{mes_especifico.upper()}-{anio_especifico}"
            ruta_archivo = f"{self.RUTA_REPORTES_ASISTENCIA}/{nombre_archivo}.xlsx"
            
            self.cargar_encabezados_asistencia_mensual(hoja, borde_celda, relleno_encabezados, fuente_negrita, fuente_titulo, alineacion_centrada)
            self.cargar_datos_asistencia_inasistencia_empleados(hoja, lista_dict_asistencia_mensual_empleados, borde_celda, alineacion_centrada, relleno_fines_semana)
            self.cargar_datos_sumatoria_asistencia_inasistencia_empleados(hoja, lista_dict_sumatoria_asistencia_inasistencia_empleados, alineacion_centrada, borde_celda)
            self.cargar_datos_dias_habiles(hoja, dias_habiles, fuente_negrita, alineacion_centrada, borde_celda, relleno_encabezados)
            self.cargar_encabezados_promedio_porcentaje_asistencia_inasistencia_empleados(hoja, fuente_negrita, alineacion_centrada, borde_celda, relleno_encabezados)
            self.cargar_datos_promedio_asistencia_inasistencia_empleados(hoja, lista_dict_promedio_asistencia_inasistencia_empleados, alineacion_centrada, borde_celda)
            self.cargar_datos_porcentaje_asistencia_inasistencia_empleados(hoja, lista_dict_porcentaje_asistencia_inasistencia_empleados, alineacion_centrada, borde_celda)
            
            libro.save(ruta_archivo)
        except Exception as error:
            raise error


if __name__ == "__main__":
    try:
        ANIO_MES = "2026-01"
        ANIO = 2026
        MES = 1
        
        reporte_asistencia_mensual_empleados = ReporteAsistenciaMensualEmpleados()
        
        datos = reporte_asistencia_mensual_empleados.cargar_datos(
            ANIO, MES
        )
        
        reporte_asistencia_mensual_empleados.exportar(datos)
        print("Se exportó el reporte de asistencia mensual de empleados correctamente")
    except Exception as error:
        print(f"ERROR AL EXPORTAR LA ASISTENCIA MENSUAL DE EMPLEADOS: {error}")
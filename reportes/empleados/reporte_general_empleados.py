from repositorios.empleados.empleado_repositorio import EmpleadoRepositorio
from repositorios.empleados.info_laboral_repositorio import InfoLaboralRepositorio
from repositorios.empleados.detalle_cargo_repositorio import DetalleCargoRepositorio
from repositorios.empleados.info_clinica_empleado_repositorio import InfoClinicaEmpleadoRepositorio

from servicios.empleados.empleado_servicio import EmpleadoServicio
from servicios.empleados.info_laboral_servicio import InfoLaboralServicio
from servicios.empleados.detalle_cargo_servicio import DetalleCargoServicio
from servicios.empleados.info_clinica_empleado_servicio import InfoClinicaEmpleadoServicio

from typing import Any, List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from reportes.reporte_base import ReporteBase
from configuraciones.configuracion import app_configuracion
from excepciones.base_datos_error import BaseDatosError
from datetime import datetime
from recursos_graficos_y_logicos.utilidades.funciones_sistema import FuncionSistema
import calendar


class ReporteGeneralEmpleados(ReporteBase):
    def __init__(self):
        self.RUTA_REPORTES_GENERALES_EMPLEADOS = app_configuracion.DIRECTORIO_REPORTES_EMPLEADOS
    
    def transformar_data_empleados(
        self,
        empleado_repositorio: EmpleadoRepositorio,
        info_laboral_repositorio: InfoLaboralRepositorio,
        detalle_cargo_repositorio: DetalleCargoRepositorio,
        info_clinica_empleado_repositorio: InfoClinicaEmpleadoRepositorio
    ) -> List[Dict]:
        lista_dict_data_personal = []
        
        empleado_servicio = EmpleadoServicio(empleado_repositorio)
        info_laboral_servicio = InfoLaboralServicio(info_laboral_repositorio)
        detalle_cargo_servicio = DetalleCargoServicio(detalle_cargo_repositorio)
        info_clinica_empleado_servicio = InfoClinicaEmpleadoServicio(info_clinica_empleado_repositorio)
        
        for empleado in empleado_servicio.obtener_todos_empleados():
            empleado_id = empleado[0]
            primer_nombre = empleado[1]
            segundo_nombre = empleado[2] if (empleado[2] is not None) else ""
            tercer_nombre = empleado[3] if (empleado[3] is not None) else ""
            apellido_paterno = empleado[4]
            apellido_materno = empleado[5] if (empleado[5] is not None) else ""
            nombre_completo = f"{primer_nombre} {segundo_nombre} {tercer_nombre} {apellido_paterno} {apellido_materno}"
            cedula = empleado[6]
            genero = empleado[10]
            fecha_nacimiento = empleado[7]
            edad = empleado[8]
            mes_nacimiento = self.cargar_mes(fecha_nacimiento)
            cargo = detalle_cargo_servicio.obtener_detalles_cargo(empleado_id)[2]
            codigo_cargo = detalle_cargo_servicio.obtener_detalles_cargo(empleado_id)[1]
            titulo_cargo = detalle_cargo_servicio.obtener_detalles_cargo(empleado_id)[5]
            funcion_cargo = detalle_cargo_servicio.obtener_detalles_cargo(empleado_id)[3]
            tipo_cargo = detalle_cargo_servicio.obtener_detalles_cargo(empleado_id)[4]
            
            programa_labor = detalle_cargo_servicio.obtener_detalles_cargo(empleado_id)[10] if (detalle_cargo_servicio.obtener_detalles_cargo(empleado_id)[10] is not None) else detalle_cargo_servicio.obtener_detalles_cargo(empleado_id)[6]
            
            fecha_ingreso_institucion = detalle_cargo_servicio.obtener_detalles_cargo(empleado_id)[7]
            tiempo_servicio = detalle_cargo_servicio.obtener_detalles_cargo(empleado_id)[9]
            estado = empleado_servicio.obtener_info_geografica_empleado(empleado_id)[1]
            municipio = empleado_servicio.obtener_info_geografica_empleado(empleado_id)[2]
            cod_dependencia_cobra = info_laboral_servicio.obtener_info_laboral_por_empleado_id(empleado_id)[4] if (info_laboral_servicio.obtener_info_laboral_por_empleado_id(empleado_id) is not None) else ""
            institucion_labora = info_laboral_servicio.obtener_info_laboral_por_empleado_id(empleado_id)[5] if (info_laboral_servicio.obtener_info_laboral_por_empleado_id(empleado_id) is not None) else ""
            
            num_telefono_principal = empleado_servicio.obtener_info_contacto_empleado(empleado_id)[1]
            num_telefono_secundario = empleado_servicio.obtener_info_contacto_empleado(empleado_id)[2] if (empleado_servicio.obtener_info_contacto_empleado(empleado_id)[2] is not None) else ""
            
            lista_telefonos = [num_telefono_principal, num_telefono_secundario]
            
            if (len(lista_telefonos) == 2):
                telefono = "\n".join(lista_telefonos)
            else:
                telefono = num_telefono_principal
            
            correo_electronico_principal = empleado_servicio.obtener_info_contacto_empleado(empleado_id)[3]
            correo_electronico_secundario = empleado_servicio.obtener_info_contacto_empleado(empleado_id)[4] if (empleado_servicio.obtener_info_contacto_empleado(empleado_id)[4] is not None) else ""
            
            lista_correo_electronico = [correo_electronico_principal, correo_electronico_secundario]
            
            if (len(lista_correo_electronico) == 2):
                correo_electronico = "\n".join(lista_correo_electronico)
            else:
                correo_electronico = correo_electronico_principal
            
            direccion_residencia = empleado_servicio.obtener_info_geografica_empleado(empleado_id)[3]
            
            lista_discapacidades_empleado = []
            
            if (info_clinica_empleado_servicio.obtener_info_clinica_por_empleado_id(empleado_id)):
                for discapacidad in info_clinica_empleado_servicio.obtener_info_clinica_por_empleado_id(empleado_id):
                    nombre_discapacidad = discapacidad[2]
                    lista_discapacidades_empleado.append(nombre_discapacidad)
            
            lista_discapacidades = lista_discapacidades_empleado
            situacion = empleado[9]
            
            if (funcion_cargo in ["DIRECTORA ENCARGADA", "SUB-DIRECTOR ENCARGADO"]):
                tipo_personal = "PERSONAL DIRECTIVO"
            elif (funcion_cargo in ["DOC. ESPECIALISTA", "DOC. AULA/FORMACIÓN", "DOC. DE AULA", "DOC. DE AULA/CULTURA"]):
                tipo_personal = "PERSONAL DOCENTE"
            elif (funcion_cargo in ["DOC. DE AULA/DEPORTE"]):
                tipo_personal = "PERSONAL DOCENTE DE DEPORTE"
            elif (funcion_cargo in ["AUXILIAR"]):
                tipo_personal = "PERSONAL AUXILIAR"
            elif (funcion_cargo in ["TERAPISTA OCUPACIONAL"]):
                tipo_personal = "PERSONAL TERAPISTA OCUPACIONAL"
            elif (funcion_cargo in ["TERAPISTA DE LENGUAJE"]):
                tipo_personal = "PERSONAL TERAPISTA DE LENGUAJE"
            elif (funcion_cargo in ["TRABAJADORA SOCIAL"]):
                tipo_personal = "PERSONAL TRABAJADORA SOCIAL"
            elif (funcion_cargo in ["SECRETARIO", "SECRETARIA"]):
                tipo_personal = "PERSONAL ADMINISTRATIVO"
            elif (funcion_cargo in ["OBRERO", "OBRERA", "COCINERA", "COCINERO", "MADRE DE LA PATRIA", "MADRE COCINERA", "PORTERO", "PORTERA"]):
                tipo_personal = "PERSONAL OBRERO"
            else:
                tipo_personal = "PERSONAL EXTRA"
            
            elemento_lista = {
                "nombre_completo": nombre_completo,
                "cedula_identidad": cedula,
                "primer_apellido": apellido_paterno,
                "segundo_apellido": apellido_materno,
                "primer_nombre": primer_nombre,
                "segundo_nombre": segundo_nombre,
                "tercer_nombre": tercer_nombre,
                "genero": "MASCULINO" if (genero == "M") else "FEMENINO",
                "fecha_nacimiento": fecha_nacimiento,
                "edad": f"{edad} años",
                "mes_nacimiento": mes_nacimiento,
                "cargo": cargo,
                "codigo_cargo": codigo_cargo,
                "titulo": titulo_cargo,
                "funcion_cargo": funcion_cargo,
                "tipo_cargo": tipo_cargo,
                "programa_o_labor": programa_labor,
                "fecha_ingreso": fecha_ingreso_institucion,
                "tiempo_servicio": f"{tiempo_servicio} años",
                "estado": estado,
                "municipio": municipio,
                "cod_dependencia_cobra": cod_dependencia_cobra,
                "institucion_labora": institucion_labora,
                "telefono": telefono,
                "correo_electronico": correo_electronico,
                "direccion_residencia": direccion_residencia,
                "discapacidades": ". ".join(lista_discapacidades),
                "situacion": situacion,
                "tipo_personal": tipo_personal
            }
            
            lista_dict_data_personal.append(elemento_lista)
            
        return lista_dict_data_personal
    
    def transformar_conteo_personal(self, detalle_cargo_repositorio: DetalleCargoRepositorio) -> List[Dict]:
        lista_dict_conteo_personal = []
        
        detalle_cargo_servicio = DetalleCargoServicio(detalle_cargo_repositorio)
        
        for elemento in detalle_cargo_servicio.conteo_empleados_por_funcion_cargo():
            funcion_cargo = elemento[0]
            cantidad_funcion_cargo = elemento[1]
            
            elemento_lista = {
                "funcion_cargo": funcion_cargo,
                "cantidad_funcion_cargo": cantidad_funcion_cargo
            }
            
            lista_dict_conteo_personal.append(elemento_lista)
        
        orden_deseado = [
            "DIRECTORA ENCARGADA",
            "SUB-DIRECTOR ENCARGADO",
            "SUBTOTAL-1",
            "DOC. ESPECIALISTA",
            "DOC. DE AULA",
            "DOC. DE AULA/DEPORTE",
            "DOC. DE AULA/CULTURA",
            "AUXILIAR",
            "TERAPISTA OCUPACIONAL",
            "TERAPISTA DE LENGUAJE",
            "TRABAJADORA SOCIAL",
            "SECRETARIO",
            "SECRETARIA",
            "OBRERO",
            "PORTERO",
            "MADRE COCINERA",
            "SUBTOTAL-2",
            "TOTAL-GENERAL"
        ]
        
        posiciones = {nombre: i for i, nombre in enumerate(orden_deseado)}
        
        lista_dict_conteo_personal.sort(key=lambda x: posiciones.get(x['funcion_cargo'], len(orden_deseado)))
        
        return lista_dict_conteo_personal
    
    def transformar_matricula_personal(self, detalles_cargo_repositorio: DetalleCargoRepositorio) -> Dict:
        detalles_cargo_servicio = DetalleCargoServicio(detalles_cargo_repositorio)
        lista_matricula_personal = detalles_cargo_servicio.conteo_matricula_empleados()
        lista_dict_matricula_personal = []
        
        for elemento in lista_matricula_personal:
            total_varones = elemento[0]
            total_hembras = elemento[1]
            total_general = elemento[2]
            
            elemento_lista = {
                "total_varones": total_varones,
                "total_hembras": total_hembras,
                "total_general": total_general
            }
            
            lista_dict_matricula_personal.append(elemento_lista)
        
        return lista_dict_matricula_personal
    
    def crear_libro(self):
        libro = Workbook()
        
        return libro
    
    def crear_hojas(self, libro):
        hoja_1 = libro.active
        hoja_1.title = "EMPLEADOS ACTIVOS"
        
        hoja_2 = libro.create_sheet("RESUMEN")
        
        return hoja_1, hoja_2
    
    def cargar_mes(self, fecha_nacimiento: datetime) -> str:
        fecha_dt = datetime.strptime(fecha_nacimiento, "%Y-%m-%d")
        
        diccionario_meses_espaniol = {
            "January": "Enero",
            "February": "Febrero",
            "March": "Marzo",
            "April": "Abril",
            "May": "Mayo",
            "June": "Junio",
            "July": "Julio",
            "August": "Agosto",
            "September": "Septiembre",
            "October": "Octubre",
            "November": "Noviembre",
            "December": "Diciembre"
        }
        
        # Obtener el nombre del mes en inglés de la fecha de nacimiento
        nombre_mes_ingles = calendar.month_name[fecha_dt.month]
        
        # Devolver el nombre del mes en español
        return diccionario_meses_espaniol.get(nombre_mes_ingles)
    
    def cargar_configuracion_excel(self, hoja):
        hoja.column_dimensions["B"].width = 50
        hoja.row_dimensions[1].height = 47
    
    def cargar_cintillo(self, hoja):
        RUTA_IMAGEN = "reportes/imagenes/CINTILLO_PEQUEÑO_TELA.png"
        
        cintillo = Image(RUTA_IMAGEN)
        cintillo.width = 50 * 7.2
        cintillo.height = 47 + 10
        
        cintillo.anchor = "A1"
        hoja.add_image(cintillo)
    
    def cargar_encabezados_info_personal(self, hoja, borde_celda, fuente_negrita, alineacion_centrada, relleno_encabezados):
        encabezados = [
            "N°", "APELLIDOS Y NOMBRES\nPERSONAL ACTIVO", "CÉDULA DE IDENTIDAD",
            "PRIMER APELLIDO", "SEGUNDO APELLIDO", "PRIMER NOMBRE", "SEGUNDO NOMBRE", "TERCER NOMBRE",
            "GÉNERO", "FECHA DE NACIMIENTO", "EDAD", "MES DE NACIMIENTO", "CARGO",
            "CÓDIGO DE CARGO", "TÍTULO", "FUNCIONES QUE CUMPLE", "TIPO DE CARGO",
            "PROGRAMA O LABOR", "FECHA DE INGRESO", "AÑOS DE SERVICIO",
            "ESTADO", "MUNICIPIO", "COD. DEPENDENCIA POR DONDE COBRA", 
            "INSTITUCIÓN DONDE LABORA", "TELÉFONO", "CORREO", "DIRECCIÓN",
            "DISCAPACIDAD", "SITUACIÓN"
        ]
        
        for i, encabezado in enumerate(encabezados, start = 1):
            letra_columna = get_column_letter(i)
            
            celda = hoja[f"{letra_columna}4"]
            
            celda.border = borde_celda
            celda.font = fuente_negrita
            celda.value = encabezado
            celda.alignment = alineacion_centrada
            celda.fill = relleno_encabezados
            
            longitud_encabezado = len(encabezado)
            
            if (encabezado == "N°"):
                hoja.column_dimensions[letra_columna].width = longitud_encabezado + 5
                hoja.row_dimensions[2].height = 47
            elif (encabezado == "TÍTULO"):
                hoja.column_dimensions[letra_columna].width = longitud_encabezado + 50
                hoja.row_dimensions[2].height = 47
            elif (encabezado == "CORREO"):
                hoja.column_dimensions[letra_columna].width = longitud_encabezado + 30
                hoja.row_dimensions[2].height = 47
            elif (encabezado == "DIRECCIÓN"):
                hoja.column_dimensions[letra_columna].width = longitud_encabezado + 70
                hoja.row_dimensions[2].height = 47
            else:
                hoja.column_dimensions[letra_columna].width = longitud_encabezado + 10
                hoja.row_dimensions[2].height = 47
    
    def cargar_data_personal(
        self,
        hoja,
        lista_dict_data_personal: List[Dict],
        fuente_seccion_personal,
        relleno_seccion_personal,
        alineacion_centrada,
        borde_celda
    ):
        fila_actual = 5
        contador_general = 1
        
        # Agrupar los empleados por tipo_personal
        personal_agrupado = {}
        
        for empleado in lista_dict_data_personal:
            tipo = empleado.get("tipo_personal", "PERSONAL EXTRA")
            
            if (tipo not in personal_agrupado):
                personal_agrupado[tipo] = []
            personal_agrupado[tipo].append(empleado)
        
        # Orden de las secciones del personal
        lista_secciones_personal = [
            "PERSONAL DIRECTIVO", "PERSONAL DOCENTE", "PERSONAL DOCENTE DE DEPORTE",
            "PERSONAL AUXILIAR", "PERSONAL TERAPISTA OCUPACIONAL", "PERSONAL TERAPISTA DE LENGUAJE",
            "PERSONAL TRABAJADORA SOCIAL", "PERSONAL ADMINISTRATIVO", "PERSONAL OBRERO", "PERSONAL EXTRA"
        ]
        
        # Pintar toda la fila con el nombre de la sección
        for seccion_personal in lista_secciones_personal:
            empleados_en_seccion = personal_agrupado.get(seccion_personal) if personal_agrupado.get(seccion_personal) is not None else ""
            
            if empleados_en_seccion:
                for columna in range(1, 30):
                    celda = hoja.cell(row = fila_actual, column = columna)
                    celda.fill = relleno_seccion_personal
                    celda.border = borde_celda
                    
                    if (columna == 2):
                        celda.value = seccion_personal
                        celda.font = fuente_seccion_personal
                        celda.alignment = alineacion_centrada
                
                hoja.row_dimensions[fila_actual].height = 25
                fila_actual += 1
            
            # Escribir en cada columna de la fila cada dato del empleado
            for empleado in empleados_en_seccion:
                datos_empleado = [
                    contador_general,
                    empleado["nombre_completo"],
                    empleado["cedula_identidad"],
                    empleado["primer_apellido"],
                    empleado["segundo_apellido"],
                    empleado["primer_nombre"],
                    empleado["segundo_nombre"],
                    empleado["tercer_nombre"],
                    empleado["genero"],
                    empleado["fecha_nacimiento"],
                    empleado["edad"],
                    empleado["mes_nacimiento"],
                    empleado["cargo"],
                    empleado["codigo_cargo"],
                    empleado["titulo"],
                    empleado["funcion_cargo"],
                    empleado["tipo_cargo"],
                    empleado["programa_o_labor"],
                    empleado["fecha_ingreso"],
                    empleado["tiempo_servicio"],
                    empleado["estado"],
                    empleado["municipio"],
                    empleado["cod_dependencia_cobra"],
                    empleado["institucion_labora"],
                    empleado["telefono"],
                    empleado["correo_electronico"],
                    empleado["direccion_residencia"],
                    empleado["discapacidades"],
                    empleado["situacion"]
                ]
                
                for indice_columna, valor in enumerate(datos_empleado, start = 1):
                    celda = hoja.cell(row = fila_actual, column = indice_columna, value = valor)
                    celda.alignment = alineacion_centrada
                    celda.border = borde_celda
                
                hoja.row_dimensions[fila_actual].height = 25
                fila_actual += 1
                contador_general += 1
    
    def cargar_encabezados_resumen_personal(self, hoja, fuente_negrita, alineacion_centrada):
        hoja["A1"] = "PERSONAL DEL TELA"
        hoja.column_dimensions["A"].width = 57
        hoja.merge_cells("A1:B1")
        hoja.cell(row = 1, column = 1).font = fuente_negrita
        hoja.cell(row = 1, column = 1).alignment = alineacion_centrada
        
        hoja["D1"] = "MATRICULA DEL TELA"
        hoja.column_dimensions["D"].width = 28.5
        hoja.column_dimensions["E"].width = 28.5
        hoja.merge_cells("D1:E1")
        hoja.cell(row = 1, column = 4).font = fuente_negrita
        hoja.cell(row = 1, column = 4).alignment = alineacion_centrada
        
        hoja["D8"] = "DATOS DEL TELA"
        hoja.merge_cells("D8:E8")
        hoja.cell(row = 8, column = 4).font = fuente_negrita
        hoja.cell(row = 8, column = 4).alignment = alineacion_centrada
    
    def cargar_conteo_personal(
        self,
        hoja,
        lista_dict_conteo_personal: List[Dict],
        borde_celda,
        fuente_negrita,
        alineacion_centrada,
        relleno_sub_totales,
        relleno_totales
    ):
        fila_actual = 3

        hoja["A2"] = "TURNO DE LA MAÑANA"
        hoja.merge_cells("A2:B2")
        hoja.cell(row = 2, column = 1).alignment = alineacion_centrada
        FuncionSistema.aplicar_borde_a_rango(hoja, "A2:B2", borde_celda)
        
        for elemento in lista_dict_conteo_personal:
            hoja.cell(row = fila_actual, column = 1, value = elemento["funcion_cargo"])
            hoja.cell(row = fila_actual, column = 1).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 1).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 2, value = elemento["cantidad_funcion_cargo"])
            hoja.cell(row = fila_actual, column = 2).font = fuente_negrita
            hoja.cell(row = fila_actual, column = 2).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 2).border = borde_celda
            
            if (elemento["funcion_cargo"] in ["SUBTOTAL-1", "SUBTOTAL-2"]):
                hoja.cell(row = fila_actual, column = 1).fill = relleno_sub_totales
                hoja.cell(row = fila_actual, column = 2).fill = relleno_sub_totales
            elif (elemento["funcion_cargo"] == "TOTAL-GENERAL"):
                hoja.cell(row = fila_actual, column = 1).fill = relleno_totales
                hoja.cell(row = fila_actual, column = 2).fill = relleno_totales
            
            fila_actual += 1
    
    def cargar_matricula_personal(
        self,
        hoja,
        lista_dict_matricula_personal: List[Dict],
        borde_celda,
        fuente_negrita,
        alineacion_centrada,
        relleno_totales
    ):
        fila_actual = 3
        columna_actual = 4
        
        elemento = lista_dict_matricula_personal[0]

        hoja["D2"] = "TURNO DE LA MAÑANA"
        hoja.merge_cells("D2:E2")
        hoja.cell(row = 2, column = columna_actual).alignment = alineacion_centrada
        FuncionSistema.aplicar_borde_a_rango(hoja, "D2:E2", borde_celda)
        
        hoja.cell(row = fila_actual, column = 4, value = "VARONES")
        hoja.cell(row = fila_actual, column = 4).alignment = alineacion_centrada
        hoja.cell(row = fila_actual, column = 4).border = borde_celda
        
        hoja.cell(row = fila_actual + 1, column = 4, value = "HEMBRAS")
        hoja.cell(row = fila_actual + 1, column = 4).alignment = alineacion_centrada
        hoja.cell(row = fila_actual + 1, column = 4).border = borde_celda
        
        hoja.cell(row = fila_actual + 2, column = 4, value = "TOTAL")
        hoja.cell(row = fila_actual + 2, column = 4).alignment = alineacion_centrada
        hoja.cell(row = fila_actual + 2, column = 4).border = borde_celda
        
        hoja.cell(row = fila_actual + 2, column = 4).fill = relleno_totales
        hoja.cell(row = fila_actual + 2, column = 5).fill = relleno_totales
        
        
        hoja.cell(row = fila_actual, column = 5, value = elemento["total_varones"])
        hoja.cell(row = fila_actual, column = 5).font = fuente_negrita
        hoja.cell(row = fila_actual, column = 5).alignment = alineacion_centrada
        hoja.cell(row = fila_actual, column = 5).border = borde_celda
        
        hoja.cell(row = fila_actual + 1, column = 5, value = elemento["total_hembras"])
        hoja.cell(row = fila_actual + 1, column = 5).font = fuente_negrita
        hoja.cell(row = fila_actual + 1, column = 5).alignment = alineacion_centrada
        hoja.cell(row = fila_actual + 1, column = 5).border = borde_celda
        
        hoja.cell(row = fila_actual + 2, column = 5, value = elemento["total_general"])
        hoja.cell(row = fila_actual + 2, column = 5).font = fuente_negrita
        hoja.cell(row = fila_actual + 2, column = 5).alignment = alineacion_centrada
        hoja.cell(row = fila_actual + 2, column = 5).border = borde_celda
    
    def cargar_datos_institucion(self, hoja, fuente_negrita, alineacion_centrada):
        INSTITUCION_DIRECCION = "Dirección: Urbanización Tricentenaria, Av. Principal, entre la Calle A y Calle 11 de diciembre. Referencias de ubicación: Frente a la cancha sector la Montañita y diagonal a Radio Kariña. Ubicación: Municipio Simón Bolivar, Parroquia Sán Cristóbal, Barcelona, Edo. Anzoátegui."

        INSTITUCION_FECHA_CREACION = "Fecha de creación: 19 de septiembre de 1979."
        INSTITUCION_TURNOS_TRABAJO = "Turnos de trabajo: Mañana."
        INSTITUCION_TELEFONOS = "Teléfonos: 0281-275-3856."
        INSTITUCION_COD_DEPENDENCIA = "Código de Dependencia: 006505188."
        INSTITUCION_COD_ESTADISTICO = "Código Estadístico: 31585."
        INSTITUCION_COD_DEA = "Código DEA: OD10820304."
        INSTITUCION_COD_POSTAL = "Código postal: 6001."
        INSTITUCION_RIF = "Rif: J-08017713-4"
        INSTITUCION_GEO_POSICION = "Geo-posición: Coordenadas; Latidud norte 101484 - oeste 64.6888."
        
        hoja["D9"] = INSTITUCION_DIRECCION
        hoja.merge_cells("D9:E15")
        hoja.cell(row = 9, column = 4).font = fuente_negrita
        hoja.cell(row = 9, column = 4).alignment = alineacion_centrada
        
        hoja["D16"] = INSTITUCION_FECHA_CREACION
        hoja.merge_cells("D16:E16")
        hoja.cell(row = 16, column = 4).font = fuente_negrita
        
        hoja["D17"] = INSTITUCION_TURNOS_TRABAJO
        hoja.merge_cells("D17:E17")
        hoja.cell(row = 17, column = 4).font = fuente_negrita
        
        hoja["D18"] = INSTITUCION_TELEFONOS
        hoja.merge_cells("D18:E18")
        hoja.cell(row = 18, column = 4).font = fuente_negrita
        
        hoja["D19"] = INSTITUCION_COD_DEPENDENCIA
        hoja.merge_cells("D19:E19")
        hoja.cell(row = 19, column = 4).font = fuente_negrita
        
        hoja["D20"] = INSTITUCION_COD_ESTADISTICO
        hoja.merge_cells("D20:E20")
        hoja.cell(row = 20, column = 4).font = fuente_negrita
        
        hoja["D21"] = INSTITUCION_COD_DEA
        hoja.merge_cells("D21:E21")
        hoja.cell(row = 21, column = 4).font = fuente_negrita
        
        hoja["D22"] = INSTITUCION_COD_POSTAL
        hoja.merge_cells("D22:E22")
        hoja.cell(row = 22, column = 4).font = fuente_negrita
        
        hoja["D23"] = INSTITUCION_RIF
        hoja.merge_cells("D23:E23")
        hoja.cell(row = 23, column = 4).font = fuente_negrita
        
        hoja["D24"] = INSTITUCION_GEO_POSICION
        hoja.merge_cells("D24:E24")
        hoja.cell(row = 24, column = 4).font = fuente_negrita
    
    def cargar_datos(
        self,
        empleado_repositorio: EmpleadoRepositorio,
        info_laboral_repositorio: InfoLaboralRepositorio,
        detalle_cargo_repositorio: DetalleCargoRepositorio,
        info_clinica_empleado_repositorio: InfoClinicaEmpleadoRepositorio
    ) -> Optional[List[Any]]:
        try:
            datos = []
            
            # Posición 0: Lista dict de la data del personal
            lista_dict_data_personal = self.transformar_data_empleados(
                empleado_repositorio,
                info_laboral_repositorio,
                detalle_cargo_repositorio,
                info_clinica_empleado_repositorio
            )
            datos.append(lista_dict_data_personal)
            
            # Posición 1: Lista dict del conteo del personal
            lista_dict_conteo_personal = self.transformar_conteo_personal(detalle_cargo_repositorio)
            datos.append(lista_dict_conteo_personal)
            
            # Posición 2: Lista dict de la matricula del personal
            lista_dict_matricula_personal = self.transformar_matricula_personal(detalle_cargo_repositorio)
            datos.append(lista_dict_matricula_personal)
            
            return datos
        except BaseDatosError as error:
            raise error
        except Exception as error:
            print(f"ERROR AL CARGAR LOS DATOS: {error}")
    
    def exportar(self, datos: List):
        try:
            self.RUTA_REPORTES_GENERALES_EMPLEADOS.mkdir(exist_ok = True)
            
            libro = self.crear_libro()
            hoja_1, hoja_2 = self.crear_hojas(libro)
            
            fuente_negrita = Font(bold = True)
            fuente_seccion_personal = Font(color = "00FF00", bold = True)
            relleno_encabezados = PatternFill(start_color = "D9D9D9", end_color = "D9D9D9", fill_type = "solid")
            relleno_seccion_personal = PatternFill(start_color = "000000", end_color = "000000", fill_type = "solid")
            relleno_sub_totales = PatternFill(start_color = "A6A6A6", end_color = "A6A6A6", fill_type = "solid")
            relleno_totales = PatternFill(start_color = "808080", end_color = "808080", fill_type = "solid")
            alineacion_centrada = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
    
            tipo_borde = Side(border_style="thin", color="000000")
            borde_celda = Border(
                left = tipo_borde,
                right = tipo_borde,
                top = tipo_borde,
                bottom = tipo_borde
            )
            
            lista_dict_data_personal = datos[0]
            lista_dict_conteo_personal = datos[1]
            lista_dict_matricula_personal = datos[2]
            
            self.cargar_configuracion_excel(hoja_1)
            self.cargar_cintillo(hoja_1)
            
            self.cargar_encabezados_info_personal(
                hoja_1,
                borde_celda,
                fuente_negrita,
                alineacion_centrada,
                relleno_encabezados
            )
            
            self.cargar_data_personal(
                hoja_1,
                lista_dict_data_personal,
                fuente_seccion_personal,
                relleno_seccion_personal,
                alineacion_centrada,
                borde_celda
            )
            
            self.cargar_encabezados_resumen_personal(
                hoja_2,
                fuente_negrita,
                alineacion_centrada
            )
            
            self.cargar_conteo_personal(
                hoja_2,
                lista_dict_conteo_personal,
                borde_celda,
                fuente_negrita,
                alineacion_centrada,
                relleno_sub_totales,
                relleno_totales
            )
            
            self.cargar_matricula_personal(
                hoja_2,
                lista_dict_matricula_personal,
                borde_celda,
                fuente_negrita,
                alineacion_centrada,
                relleno_totales
            )
            
            self.cargar_datos_institucion(
                hoja_2,
                fuente_negrita,
                alineacion_centrada
            )
            
            fecha_actual = int(datetime.now().date().year)
            nombre_archivo = f"REPORTE DE PERSONAL ACTIVO - {fecha_actual}"
            
            libro.save(f"{self.RUTA_REPORTES_GENERALES_EMPLEADOS}/{nombre_archivo}.xlsx")
        except BaseDatosError as error:
            raise error
        except Exception as error:
            raise error


if __name__ == "__main__":
    reporte_general_empleados = ReporteGeneralEmpleados()
    
    empleado_repositorio = EmpleadoRepositorio()
    info_laboral_repositorio = InfoLaboralRepositorio()
    detalle_cargo_repositorio = DetalleCargoRepositorio()
    info_clinica_empleado_repositorio = InfoClinicaEmpleadoRepositorio()
    
    try:
        datos = reporte_general_empleados.cargar_datos(
            empleado_repositorio,
            info_laboral_repositorio,
            detalle_cargo_repositorio,
            info_clinica_empleado_repositorio
        )
    
        reporte_general_empleados.exportar(datos)
    except BaseDatosError as error:
        print(error)

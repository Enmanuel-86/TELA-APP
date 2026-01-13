import calendar
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Tuple, Dict, Optional, Any
from repositorios.alumnos.asistencia_alumno_repositorio import AsistenciaAlumnoRepositorio
from repositorios.especialidades.especialidad_repositorio import EspecialidadRepositorio
from servicios.alumnos.asistencia_alumno_servicio import AsistenciaAlumnoServicio
from servicios.especialidades.especialidad_servicio import EspecialidadServicio
from configuraciones.configuracion import app_configuracion
from reportes.reporte_base import ReporteBase
from excepciones.base_datos_error import BaseDatosError
from recursos_graficos_y_logicos.utilidades.funciones_sistema import FuncionSistema


class ReporteAsistenciaMensualAlumnos(ReporteBase):
    def __init__(self):
        self.RUTA_REPORTES_ASISTENCIA = app_configuracion.DIRECTORIO_REPORTES_ALUMNOS
    
    def transformar_lista_asistencia_inasistencia_alumnos(self, lista_asistencia_inasistencia_alumnos: List, lista_dia_mes_con_semana: List, anio: int, mes: int) -> Optional[List[Dict]]:
        lista_dict_asistencia_inasistencia_alumnos = []
    
        for dia_numero_str, dia_semana in lista_dia_mes_con_semana:
            dia_numero = int(dia_numero_str)
            fecha_actual = datetime.datetime(anio, mes, dia_numero).date()
            
            registro_encontrado = None
            for registro_bd in lista_asistencia_inasistencia_alumnos:
                if registro_bd[1] == str(fecha_actual):
                    registro_encontrado = registro_bd
                    break
            
            fila_tabla = {
                "Dia_Numero": dia_numero_str,
                "Dia_Semana": dia_semana,
                "Es_Feriado": None,
                "Varones_Presentes": 0,
                "Hembras_Presentes": 0,
                "Total_Presentes": 0,
                "Varones_Ausentes": 0,
                "Hembras_Ausentes": 0,
                "Total_Ausentes": 0
            }
            
            if registro_encontrado:
                fila_tabla["Es_Feriado"] = registro_encontrado[2]
                fila_tabla["Varones_Presentes"] = registro_encontrado[3]
                fila_tabla["Hembras_Presentes"] = registro_encontrado[4]
                fila_tabla["Total_Presentes"] = registro_encontrado[5]
                fila_tabla["Varones_Ausentes"] = registro_encontrado[6]
                fila_tabla["Hembras_Ausentes"] = registro_encontrado[7]
                fila_tabla["Total_Ausentes"] = registro_encontrado[8]
            else:
                if dia_semana in ["Sábado", "Domingo"]:
                    fila_tabla["Es_Feriado"] = "Fin de Semana"
            
            lista_dict_asistencia_inasistencia_alumnos.append(fila_tabla)
        return lista_dict_asistencia_inasistencia_alumnos
    
    def transformar_lista_totales_asistencia_inasistencia_alumnos(self, tupla_asistencia_inasistencia_alumnos: Tuple) -> Optional[List[Dict]]:
        lista_dict_asistencia_inasistencia_alumnos = []
    
        sumatorio_varones_presentes = tupla_asistencia_inasistencia_alumnos[3]
        sumatoria_hembras_presentes = tupla_asistencia_inasistencia_alumnos[4]
        sumatoria_general_presentes = tupla_asistencia_inasistencia_alumnos[5]
            
        sumatoria_varones_ausentes = tupla_asistencia_inasistencia_alumnos[6]
        sumatoria_hembras_ausentes = tupla_asistencia_inasistencia_alumnos[7]
        sumatoria_general_ausentes = tupla_asistencia_inasistencia_alumnos[8]
            
        fila_tabla = {
            "Sumatoria_Varones_Presentes": sumatorio_varones_presentes,
            "Sumatoria_Hembras_Presentes": sumatoria_hembras_presentes,
            "Sumatoria_General_Presentes": sumatoria_general_presentes,
            "Sumatoria_Varones_Ausentes": sumatoria_varones_ausentes,
            "Sumatoria_Hembras_Ausentes": sumatoria_hembras_ausentes,
            "Sumatoria_General_Ausentes": sumatoria_general_ausentes
        }
            
        lista_dict_asistencia_inasistencia_alumnos.append(fila_tabla)
        
        return lista_dict_asistencia_inasistencia_alumnos
    
    def transformar_promedio_asistencia_inasistencia_alumnos(self, tupla_promedio_asistencia_inasistencia_alumnos: Tuple) -> Optional[List[Dict]]:
        lista_dict_promedio_asistencia_inasistencia_alumnos = []
    
        promedio_varones_presentes = tupla_promedio_asistencia_inasistencia_alumnos[3]
        promedio_hembras_presentes = tupla_promedio_asistencia_inasistencia_alumnos[4]
        promedio_general_presentes = tupla_promedio_asistencia_inasistencia_alumnos[5]
        
        promedio_varones_ausentes = tupla_promedio_asistencia_inasistencia_alumnos[6]
        promedio_hembras_ausentes = tupla_promedio_asistencia_inasistencia_alumnos[7]
        promedio_general_ausentes = tupla_promedio_asistencia_inasistencia_alumnos[8]
        
        fila_tabla = {
            "Promedio_Varones_Presentes": promedio_varones_presentes,
            "Promedio_Hembras_Presentes": promedio_hembras_presentes,
            "Promedio_General_Presentes": promedio_general_presentes,
            "Promedio_Varones_Ausentes": promedio_varones_ausentes,
            "Promedio_Hembras_Ausentes": promedio_hembras_ausentes,
            "Promedio_General_Ausentes": promedio_general_ausentes
        }
        
        lista_dict_promedio_asistencia_inasistencia_alumnos.append(fila_tabla)
        
        return lista_dict_promedio_asistencia_inasistencia_alumnos
    
    def transformar_porcentaje_asistencia_inasistencia_alumnos(self, tupla_porcentaje_asistencia_inasistencia_alumnos: Tuple) -> Optional[List[Dict]]:
        lista_dict_porcentaje_asistencia_inasistencia_alumnos = []
    
        porcentaje_varones_presentes = tupla_porcentaje_asistencia_inasistencia_alumnos[3]
        porcentaje_hembras_presentes = tupla_porcentaje_asistencia_inasistencia_alumnos[4]
        porcentaje_general_presentes = tupla_porcentaje_asistencia_inasistencia_alumnos[5]
        
        porcentaje_varones_ausentes = tupla_porcentaje_asistencia_inasistencia_alumnos[6]
        porcentaje_hembras_ausentes = tupla_porcentaje_asistencia_inasistencia_alumnos[7]
        porcentaje_general_ausentes = tupla_porcentaje_asistencia_inasistencia_alumnos[8]
        
        fila_tabla = {
            "Porcentaje_Varones_Presentes": porcentaje_varones_presentes,
            "Porcentaje_Hembras_Presentes": porcentaje_hembras_presentes,
            "Porcentaje_General_Presentes": porcentaje_general_presentes,
            "Porcentaje_Varones_Ausentes": porcentaje_varones_ausentes,
            "Porcentaje_Hembras_Ausentes": porcentaje_hembras_ausentes,
            "Porcentaje_General_Ausentes": porcentaje_general_ausentes
        }
        
        lista_dict_porcentaje_asistencia_inasistencia_alumnos.append(fila_tabla)
        
        return lista_dict_porcentaje_asistencia_inasistencia_alumnos
    
    def transformar_matricula_completa_alumnos(self, lista_matricula_alumnos: List) -> Optional[List[Dict]]:
        lista_dict_matricula_completa_alumnos = []
    
        for cantidad in lista_matricula_alumnos:
            total_varones = cantidad[0]
            total_hembras = cantidad[1]
            total_general = cantidad[2]
            
            fila_tabla = {
                "Total_Varones": total_varones,
                "Total_Hembras": total_hembras,
                "Total_General": total_general
            }
            
            lista_dict_matricula_completa_alumnos.append(fila_tabla)
        
        return lista_dict_matricula_completa_alumnos
    
    def cargar_dia_mes_con_dia_semana(self, anio: int, mes: int) -> List[Tuple]:
        num_dias = calendar.monthrange(anio, mes)[1]
        dias_con_semana = []
        diccionario_dias_semana_espaniol = {
            "Monday": "Lunes",
            "Tuesday": "Martes",
            "Wednesday": "Miercoles",
            "Thursday": "Jueves",
            "Friday": "Viernes",
            "Saturday": "Sábado",
            "Sunday": "Domingo"
        }
        
        for dia in range(1, num_dias + 1):
            fecha = datetime.date(anio, mes, dia)
            fecha_formateada = fecha.strftime("%d")
            
            dia_semana_ingles = calendar.day_name[fecha.weekday()]
            dia_semana_espaniol = diccionario_dias_semana_espaniol[dia_semana_ingles]
            
            dias_con_semana.append((fecha_formateada, dia_semana_espaniol))
        return dias_con_semana
    
    def cargar_meses_y_anio(self, anio: int) -> List[str]:
        meses_espaniol_y_anio = []
        meses_ingles = list(calendar.month_name)
        
        diccionario_meses_espaniol = {
            "January": "Enero",
            "February": "Febrero",
            "March": "Marzo",
            "April": "Abril",
            "May": "Mayo",
            "June": "Junio",
            "July": "Julio",
            "August": "Agosto",
            "September": "Septiembre",
            "October": "Octubre",
            "November": "Noviembre",
            "December": "Diciembre"
        }
        
        for mes in range(1, 13):
            mes_ingles = meses_ingles[mes]
            mes_espaniol = diccionario_meses_espaniol[mes_ingles]
            meses_espaniol_y_anio.append(mes_espaniol)
        meses_espaniol_y_anio.append(anio)
        
        return meses_espaniol_y_anio
    
    def cargar_datos(
            self, asistencia_alumno_repositorio: AsistenciaAlumnoRepositorio,
            especialidad_repositorio: EspecialidadRepositorio,
            especialidad_id: int, anio: int, mes: int
        ) -> Optional[List[Any]]:
        try:
            datos = []
            
            asistencia_alumnos_servicio = AsistenciaAlumnoServicio(asistencia_alumno_repositorio)
            especialidades_servicio = EspecialidadServicio(especialidad_repositorio)
            
            if (mes > 9):
                anio_mes = f"{anio}-{mes}"
            else:
                anio_mes = f"{anio}-0{mes}"
            
            # Posición 0: Día del mes con el día de semana
            lista_dia_mes_con_dia_semana = self.cargar_dia_mes_con_dia_semana(anio, mes)
            datos.append(lista_dia_mes_con_dia_semana)
            
            # Posición 1: Mes específico
            mes_especifico = self.cargar_meses_y_anio(anio)[(mes - 1)]
            datos.append(mes_especifico)
            
            # Posición 2: Año específico
            anio_especifico = self.cargar_meses_y_anio(anio)[12]
            datos.append(anio_especifico)
            
            # Posición 3: Nombre de la Especialidad
            nombre_especialidad = especialidades_servicio.obtener_especialidad_por_id(especialidad_id)[1]
            datos.append(nombre_especialidad)
            
            # Posición 4: Lista Dict de la asistencia mensual de alumnos
            if not(asistencia_alumnos_servicio.obtener_asistencia_por_especialidad_y_anio_mes(especialidad_id, anio_mes)):
                raise BaseDatosError("ASISTENCIA_ALUMNOS_NO_EXISTE", "No existen registros de asistencia de alumnos en esta especialidad, año y mes.")
            
            lista_asistencia_inasistencia_alumnos = asistencia_alumnos_servicio.obtener_asistencia_por_especialidad_y_anio_mes(especialidad_id, anio_mes)
            lista_dict_asistencia_inasistencia_alumnos = self.transformar_lista_asistencia_inasistencia_alumnos(lista_asistencia_inasistencia_alumnos, lista_dia_mes_con_dia_semana, anio, mes)
            datos.append(lista_dict_asistencia_inasistencia_alumnos)
            
            # Posición 5: Lista Dict de la sumatoria de asistencias e inasistencias de alumnos mensual
            tupla_sumatoria_asistencia_inasistencias_alumnos = asistencia_alumnos_servicio.obtener_sumatoria_asistencias_inasistencias(especialidad_id, anio_mes)
            lista_dict_sumatoria_asistencia_inasistencia_alumnos = self.transformar_lista_totales_asistencia_inasistencia_alumnos(tupla_sumatoria_asistencia_inasistencias_alumnos)
            datos.append(lista_dict_sumatoria_asistencia_inasistencia_alumnos)
            
            # Posición 6: Lista Dict de la matricula completa de alumnos
            tupla_matricula_completa_alumnos = asistencia_alumnos_servicio.obtener_matricula_completa_alumnos(especialidad_id)
            lista_matricula_completa_alumnos = [
                (tupla_matricula_completa_alumnos[2], tupla_matricula_completa_alumnos[3], tupla_matricula_completa_alumnos[4]),
                (tupla_matricula_completa_alumnos[5], tupla_matricula_completa_alumnos[6], tupla_matricula_completa_alumnos[7]),
                (tupla_matricula_completa_alumnos[8], tupla_matricula_completa_alumnos[9], tupla_matricula_completa_alumnos[10]),
                (tupla_matricula_completa_alumnos[11], tupla_matricula_completa_alumnos[12], tupla_matricula_completa_alumnos[13]),
                (tupla_matricula_completa_alumnos[14], tupla_matricula_completa_alumnos[15], tupla_matricula_completa_alumnos[16])
            ]
            lista_dict_matricula_alumnos = self.transformar_matricula_completa_alumnos(lista_matricula_completa_alumnos)
            datos.append(lista_dict_matricula_alumnos)
            
            # Posición 7: Lista Dict del promedio de asistencias e inasistencias de alumnos mensual
            tupla_promedio_asistencia_inasistencia_alumnos = asistencia_alumnos_servicio.obtener_promedio_asistencias_inasistencias(especialidad_id, anio_mes)
            lista_dict_promedio_asistencia_inasistencia_alumnos = self.transformar_promedio_asistencia_inasistencia_alumnos(tupla_promedio_asistencia_inasistencia_alumnos)
            datos.append(lista_dict_promedio_asistencia_inasistencia_alumnos)
            
            # Posición 8: Lista Dict del porcentaje de asistencias e inasistencias de alumnos mensual
            tupla_porcentaje_asistencia_inasistencia_alumnos = asistencia_alumnos_servicio.obtener_porcentaje_asistencias_inasistencias(especialidad_id, anio_mes)
            lista_dict_porcentaje_asistencia_inasistencia_alumnos = self.transformar_porcentaje_asistencia_inasistencia_alumnos(tupla_porcentaje_asistencia_inasistencia_alumnos)
            datos.append(lista_dict_porcentaje_asistencia_inasistencia_alumnos)
            
            # Posición 9: Total de Días hábiles
            dias_habiles = asistencia_alumnos_servicio.obtener_dias_habiles(especialidad_id, anio_mes)
            datos.append(dias_habiles)
            
            return datos
        except BaseDatosError as error:
            raise error
        except Exception as error:
            print(f"ERROR AL CARGAR LOS DATOS: {error}")
    
    def crear_libro_y_hoja(self):
        libro = Workbook()
        hoja = libro.active
        
        return libro, hoja
    
    def cargar_configuraciones_excel(self):
        fuente_negrita = Font(bold = True)
        alineacion_centrada = Alignment(horizontal = "center", vertical = "center")
        fuente_titulo = Font(size = 20, bold = True)
        relleno_encabezados = PatternFill(start_color = "92D050", end_color = "92D050", fill_type = "solid")
        relleno_fines_semana = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")

        tipo_borde = Side(border_style="thin", color="000000")
        borde_celda = Border(
            left = tipo_borde,
            right = tipo_borde,
            top = tipo_borde,
            bottom = tipo_borde
        )
        
        return fuente_negrita, alineacion_centrada, fuente_titulo, relleno_encabezados, relleno_fines_semana, borde_celda
    
    def cargar_encabezados_asistencia_mensual(self, hoja, nombre_especialidad: str, borde_celda, relleno_encabezados, fuente_negrita, fuente_titulo, alineacion_centrada):
        hoja["B1"] = f"{nombre_especialidad.upper()}"
        hoja["A3"] = "Fecha"
        hoja["B3"] = "Día"
        hoja["C3"] = "Asistencia"
        hoja["F3"] = "Inasistencia"

        hoja["B1"].border = borde_celda
        hoja["A3"].border = borde_celda
        hoja["B3"].border = borde_celda
        hoja["C3"].border = borde_celda
        hoja["F3"].border = borde_celda

        hoja["B1"].fill = relleno_encabezados
        hoja["A3"].fill = relleno_encabezados
        hoja["B3"].fill = relleno_encabezados
        hoja["C3"].fill = relleno_encabezados
        hoja["F3"].fill = relleno_encabezados

        hoja.merge_cells("B1:E1")
        hoja.merge_cells("A3:A4")
        hoja.merge_cells("B3:B4")
        hoja.merge_cells("C3:E3")
        hoja.merge_cells("F3:H3")

        hoja["C4"] = "V"
        hoja["D4"] = "H"
        hoja["E4"] = "T"

        hoja["C4"].border = borde_celda
        hoja["D4"].border = borde_celda
        hoja["E4"].border = borde_celda

        hoja["F4"] = "V"
        hoja["G4"] = "H"
        hoja["H4"] = "T"

        hoja["F4"].border = borde_celda
        hoja["G4"].border = borde_celda
        hoja["H4"].border = borde_celda


        for num_fila in range(1, 5):
            for num_columna in range(1, 9):
                celda = hoja.cell(row = num_fila, column = num_columna)
                
                celda.font = fuente_negrita
                celda.alignment = alineacion_centrada
                
        hoja["B1"].font = fuente_titulo
        hoja["B1"].border = borde_celda

        for columna in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            hoja.column_dimensions[columna].width = 12
    
    def cargar_datos_asistencia_inasistencia_alumnos(self, hoja, lista_dict_asistencia_mensual_alumnos: List[Dict], borde_celda, alineacion_centrada, relleno_fines_semana):
        fila_actual = 5

        for fila_data in lista_dict_asistencia_mensual_alumnos:
            # Escribimos el número de día
            hoja.cell(row = fila_actual, column = 1, value = fila_data["Dia_Numero"])
            hoja.cell(row = fila_actual, column = 1).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 1).border = borde_celda
            
            # Escribimos el día de semana
            hoja.cell(row = fila_actual, column = 2, value = fila_data["Dia_Semana"])
            hoja.cell(row = fila_actual, column = 2).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 2).border = borde_celda
            
            # Manejo de Feriados
            if (fila_data["Es_Feriado"]) and not(fila_data["Es_Feriado"] == "Fin de Semana"):
                rango = f"C{fila_actual}:H{fila_actual}"
                hoja.merge_cells(start_row = fila_actual, start_column = 3, end_row = fila_actual, end_column = 8)
                # Si es feriado o fin de semana escribir el contenido y combinar celdas
                hoja.cell(row = fila_actual, column = 3, value = fila_data["Es_Feriado"])
                hoja.cell(row = fila_actual, column = 3).alignment = alineacion_centrada
                #hoja.cell(row = fila_actual, column = 3).border = borde_celda
                
                FuncionSistema.aplicar_borde_a_rango(hoja,rango, borde_celda)
                
            # Si no es Feriado ni fin de semana, escribir la cantidad de asistencia
            elif fila_data["Es_Feriado"] == "Fin de Semana":
                rango = f"C{fila_actual}:H{fila_actual}"
                hoja.merge_cells(start_row = fila_actual, start_column = 3, end_row = fila_actual, end_column = 8)
                hoja.cell(row = fila_actual, column = 3, value = "").fill = relleno_fines_semana
                hoja.cell(row = fila_actual, column = 3).border = borde_celda
                
                hoja.cell(row = fila_actual, column = 3).alignment = alineacion_centrada
                FuncionSistema.aplicar_borde_a_rango(hoja, rango, borde_celda)
            else:
                hoja.cell(row = fila_actual, column = 3, value = fila_data["Varones_Presentes"])
                hoja.cell(row = fila_actual, column = 3).alignment = alineacion_centrada
                hoja.cell(row = fila_actual, column = 3).border = borde_celda
                
                hoja.cell(row = fila_actual, column = 4, value = fila_data["Hembras_Presentes"])
                hoja.cell(row = fila_actual, column = 4).alignment = alineacion_centrada
                hoja.cell(row = fila_actual, column = 4).border = borde_celda
                
                hoja.cell(row = fila_actual, column = 5, value = fila_data["Total_Presentes"])
                hoja.cell(row = fila_actual, column = 5).alignment = alineacion_centrada
                hoja.cell(row = fila_actual, column = 5).border = borde_celda
                
                hoja.cell(row = fila_actual, column = 6, value = fila_data["Varones_Ausentes"])
                hoja.cell(row = fila_actual, column = 6).alignment = alineacion_centrada
                hoja.cell(row = fila_actual, column = 6).border = borde_celda
                
                hoja.cell(row = fila_actual, column = 7, value = fila_data["Hembras_Ausentes"])
                hoja.cell(row = fila_actual, column = 7).alignment = alineacion_centrada
                hoja.cell(row = fila_actual, column = 7).border = borde_celda
                
                hoja.cell(row = fila_actual, column = 8, value = fila_data["Total_Ausentes"])
                hoja.cell(row = fila_actual, column = 8).alignment = alineacion_centrada
                hoja.cell(row = fila_actual, column = 8).border = borde_celda
            
            fila_actual += 1
    
    def cargar_datos_sumatoria_asistencia_inasistencia_alumnos(self, hoja, lista_dict_sumatoria_asistencia_inasistencia_alumnos: List[Dict], alineacion_centrada, borde_celda):
        ultima_fila = hoja.max_row
        siguiente_fila_disponible = ultima_fila + 1

        hoja[f"A{siguiente_fila_disponible}"] = "Total"

        hoja[f"A{siguiente_fila_disponible}"].alignment = alineacion_centrada
        hoja[f"A{siguiente_fila_disponible}"].border = borde_celda
        hoja.row_dimensions[siguiente_fila_disponible].height = 25


        for fila_data in lista_dict_sumatoria_asistencia_inasistencia_alumnos:
            hoja.cell(row = siguiente_fila_disponible, column = 3, value = fila_data["Sumatoria_Varones_Presentes"])
            hoja.cell(row = siguiente_fila_disponible, column = 3).alignment = alineacion_centrada
            hoja.cell(row = siguiente_fila_disponible, column = 3).border = borde_celda
            
            hoja.cell(row = siguiente_fila_disponible, column = 4, value = fila_data["Sumatoria_Hembras_Presentes"])
            hoja.cell(row = siguiente_fila_disponible, column = 4).alignment = alineacion_centrada
            hoja.cell(row = siguiente_fila_disponible, column = 4).border = borde_celda
            
            hoja.cell(row = siguiente_fila_disponible, column = 5, value = fila_data["Sumatoria_General_Presentes"])
            hoja.cell(row = siguiente_fila_disponible, column = 5).alignment = alineacion_centrada
            hoja.cell(row = siguiente_fila_disponible, column = 5).border = borde_celda
            
            hoja.cell(row = siguiente_fila_disponible, column = 6, value = fila_data["Sumatoria_Varones_Ausentes"])
            hoja.cell(row = siguiente_fila_disponible, column = 6).alignment = alineacion_centrada
            hoja.cell(row = siguiente_fila_disponible, column = 6).border = borde_celda
            
            hoja.cell(row = siguiente_fila_disponible, column = 7, value = fila_data["Sumatoria_Hembras_Ausentes"])
            hoja.cell(row = siguiente_fila_disponible, column = 7).alignment = alineacion_centrada
            hoja.cell(row = siguiente_fila_disponible, column = 7).border = borde_celda
            
            hoja.cell(row = siguiente_fila_disponible, column = 8, value = fila_data["Sumatoria_General_Ausentes"])
            hoja.cell(row = siguiente_fila_disponible, column = 8).alignment = alineacion_centrada
            hoja.cell(row = siguiente_fila_disponible, column = 8).border = borde_celda


        hoja.merge_cells(f"A{siguiente_fila_disponible}:B{siguiente_fila_disponible}")
        
        rango = f"A{siguiente_fila_disponible}:B{siguiente_fila_disponible}"
        FuncionSistema.aplicar_borde_a_rango(hoja, rango, borde_celda)
    
    def cargar_datos_matricula_alumnos(self, hoja, lista_dict_matricula_completa_alumnos: List[Dict], fuente_negrita, alineacion_centrada, borde_celda, relleno_encabezados):
        hoja["J8"] = "Matrícula"
        hoja["J9"] = "Inicial"
        hoja["J10"] = "Ingresos"
        hoja["J11"] = "Egresos"
        hoja["J12"] = "Rotaciones/Traslado"
        hoja["J13"] = "Final"

        hoja["M8"] = "V"
        hoja["N8"] = "H"
        hoja["O8"] = "T"

        hoja["J8"].font = fuente_negrita
        hoja["M8"].font = fuente_negrita
        hoja["N8"].font = fuente_negrita
        hoja["O8"].font = fuente_negrita

        hoja["J8"].alignment = alineacion_centrada
        hoja["M8"].alignment = alineacion_centrada
        hoja["N8"].alignment = alineacion_centrada
        hoja["O8"].alignment = alineacion_centrada

        hoja["M8"].border = borde_celda
        hoja["N8"].border = borde_celda
        hoja["O8"].border = borde_celda


        hoja["J8"].fill = relleno_encabezados
        hoja["M8"].fill = relleno_encabezados
        hoja["N8"].fill = relleno_encabezados
        hoja["O8"].fill = relleno_encabezados

        hoja.merge_cells("J8:L8")
        hoja.merge_cells("J9:L9")
        hoja.merge_cells("J10:L10")
        hoja.merge_cells("J11:L11")
        hoja.merge_cells("J12:L12")
        hoja.merge_cells("J13:L13")
        
        rangos = ["J8:L8", "J9:L9", "J10:L10", "J11:L11", "J12:L12", "J13:L13"]
        
        for rango in rangos:
            FuncionSistema.aplicar_borde_a_rango(hoja, rango, borde_celda)


        fila_actual = 9

        for fila_data in lista_dict_matricula_completa_alumnos:
            hoja.cell(row = fila_actual, column = 13, value = fila_data["Total_Varones"])
            hoja.cell(row = fila_actual, column = 13).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 13).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 14, value = fila_data["Total_Hembras"])
            hoja.cell(row = fila_actual, column = 14).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 14).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 15, value = fila_data["Total_General"])
            hoja.cell(row = fila_actual, column = 15).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 15).border = borde_celda
            
            fila_actual += 1
    
    def cargar_datos_dias_habiles(self, hoja, dias_habiles: int, fuente_negrita, alineacion_centrada, borde_celda, relleno_encabezados):
        hoja["J22"] = "Días Hábiles"
        hoja["J22"].font = fuente_negrita
        hoja["J22"].alignment = alineacion_centrada
        hoja["J22"].border = borde_celda
        hoja["J22"].fill = relleno_encabezados

        hoja.merge_cells("J22:K22")

        hoja["L22"] = dias_habiles
        hoja["L22"].alignment = alineacion_centrada
        hoja["L22"].border = borde_celda
    
    def cargar_encabezados_promedio_porcentaje_asistencia_inasistencia_alumnos(self, hoja, fuente_negrita, alineacion_centrada, borde_celda, relleno_encabezados):
        hoja["L27"] = "Asistencia"
        hoja["L27"].font = fuente_negrita
        hoja["L27"].alignment = alineacion_centrada
        hoja["L27"].border = borde_celda
        hoja["L27"].fill = relleno_encabezados

        hoja["L28"] = "V"
        hoja["M28"] = "H"
        hoja["N28"] = "T"
        hoja["L28"].font = fuente_negrita
        hoja["M28"].font = fuente_negrita
        hoja["N28"].font = fuente_negrita
        hoja["L28"].alignment = alineacion_centrada
        hoja["M28"].alignment = alineacion_centrada
        hoja["N28"].alignment = alineacion_centrada
        hoja["L28"].border = borde_celda
        hoja["M28"].border = borde_celda
        hoja["N28"].border = borde_celda


        hoja["P27"] = "Inasistencia"
        hoja["P27"].font = fuente_negrita
        hoja["P27"].alignment = alineacion_centrada
        hoja["P27"].border = borde_celda
        hoja["P27"].fill = relleno_encabezados

        hoja["P28"] = "V"
        hoja["Q28"] = "H"
        hoja["R28"] = "T"
        hoja["P28"].font = fuente_negrita
        hoja["Q28"].font = fuente_negrita
        hoja["R28"].font = fuente_negrita
        hoja["P28"].alignment = alineacion_centrada
        hoja["Q28"].alignment = alineacion_centrada
        hoja["R28"].alignment = alineacion_centrada
        hoja["P28"].border = borde_celda
        hoja["Q28"].border = borde_celda
        hoja["R28"].border = borde_celda


        hoja.merge_cells("L27:N27")
        hoja.merge_cells("P27:R27")


        hoja["J28"] = "Indicadores"
        hoja["J28"].font = fuente_negrita
        hoja["J28"].border = borde_celda
        hoja["J28"].fill = relleno_encabezados


        hoja["J29"] = "Promedio"
        hoja["J29"].font = fuente_negrita
        hoja["J29"].border = borde_celda
        hoja["J29"].fill = relleno_encabezados

        hoja["J30"] = "Porcentaje"
        hoja["J30"].font = fuente_negrita
        hoja["J30"].border = borde_celda
        hoja["J30"].fill = relleno_encabezados


        hoja.merge_cells("J28:K28")
        hoja.merge_cells("J29:K29")
        hoja.merge_cells("J30:K30")
        
        rangos = ["L27:N27", "P27:R27","J28:K28", "J29:K29", "J30:K30"]
        
        for rango in rangos:
            FuncionSistema.aplicar_borde_a_rango(hoja, rango, borde_celda)
    
    def cargar_datos_promedio_asistencia_inasistencia_alumnos(self, hoja, lista_dict_promedio_asistencia_inasistencia_alumnos: List[Dict], alineacion_centrada, borde_celda):
        fila_actual = 29

        for fila_data in lista_dict_promedio_asistencia_inasistencia_alumnos:
            hoja.cell(row = fila_actual, column = 12, value = fila_data["Promedio_Varones_Presentes"])
            hoja.cell(row = fila_actual, column = 12).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 12).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 13, value = fila_data["Promedio_Hembras_Presentes"])
            hoja.cell(row = fila_actual, column = 13).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 13).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 14, value = fila_data["Promedio_General_Presentes"])
            hoja.cell(row = fila_actual, column = 14).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 14).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 16, value = fila_data["Promedio_Varones_Ausentes"])
            hoja.cell(row = fila_actual, column = 16).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 16).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 17, value = fila_data["Promedio_Hembras_Ausentes"])
            hoja.cell(row = fila_actual, column = 17).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 17).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 18, value = fila_data["Promedio_General_Ausentes"])
            hoja.cell(row = fila_actual, column = 18).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 18).border = borde_celda
            
            fila_actual += 1
    
    def cargar_datos_porcentaje_asistencia_inasistencia_alumnos(self, hoja, lista_dict_porcentaje_asistencia_inasistencia_alumnos: List[Dict], alineacion_centrada, borde_celda):
        fila_actual = 30

        for fila_data in lista_dict_porcentaje_asistencia_inasistencia_alumnos:
            hoja.cell(row = fila_actual, column = 12, value = fila_data["Porcentaje_Varones_Presentes"])
            hoja.cell(row = fila_actual, column = 12).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 12).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 13, value = fila_data["Porcentaje_Hembras_Presentes"])
            hoja.cell(row = fila_actual, column = 13).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 13).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 14, value = fila_data["Porcentaje_General_Presentes"])
            hoja.cell(row = fila_actual, column = 14).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 14).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 16, value = fila_data["Porcentaje_Varones_Ausentes"])
            hoja.cell(row = fila_actual, column = 16).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 16).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 17, value = fila_data["Porcentaje_Hembras_Ausentes"])
            hoja.cell(row = fila_actual, column = 17).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 17).border = borde_celda
            
            hoja.cell(row = fila_actual, column = 18, value = fila_data["Porcentaje_General_Ausentes"])
            hoja.cell(row = fila_actual, column = 18).alignment = alineacion_centrada
            hoja.cell(row = fila_actual, column = 18).border = borde_celda
            
            fila_actual += 1
    
    def exportar(self, datos: List):
        try:
            self.RUTA_REPORTES_ASISTENCIA.mkdir(exist_ok = True)
                
            fuente_negrita = self.cargar_configuraciones_excel()[0]
            alineacion_centrada = self.cargar_configuraciones_excel()[1]
            fuente_titulo = self.cargar_configuraciones_excel()[2]
            relleno_encabezados = self.cargar_configuraciones_excel()[3]
            relleno_fines_semana = self.cargar_configuraciones_excel()[4]
            borde_celda = self.cargar_configuraciones_excel()[5]
                
            mes_especifico = datos[1]
            anio_especifico = datos[2]
            nombre_especialidad = datos[3]
            lista_dict_asistencia_mensual_alumnos = datos[4]
            lista_dict_sumatoria_asistencia_inasistencia_alumnos = datos[5]
            lista_dict_matricula_completa_alumnos = datos[6]
            lista_dict_promedio_asistencia_inasistencia_alumnos = datos[7]
            lista_dict_porcentaje_asistencia_inasistencia_alumnos = datos[8]
            dias_habiles = datos[9]
                
            libro, hoja = self.crear_libro_y_hoja()
                
            hoja.title = f"ASISTENCIA {mes_especifico.upper()}-{anio_especifico}"
            nombre_archivo = f"REPORTE_ASISTENCIA_ALUMNOS_{nombre_especialidad.upper()}_{mes_especifico.upper()}-{anio_especifico}"
            ruta_archivo = f"{self.RUTA_REPORTES_ASISTENCIA}/{nombre_archivo}.xlsx"
                
            self.cargar_encabezados_asistencia_mensual(hoja, nombre_especialidad, borde_celda, relleno_encabezados, fuente_negrita, fuente_titulo, alineacion_centrada)
            self.cargar_datos_asistencia_inasistencia_alumnos(hoja, lista_dict_asistencia_mensual_alumnos, borde_celda, alineacion_centrada, relleno_fines_semana)
            self.cargar_datos_sumatoria_asistencia_inasistencia_alumnos(hoja, lista_dict_sumatoria_asistencia_inasistencia_alumnos, alineacion_centrada, borde_celda)
            self.cargar_datos_matricula_alumnos(hoja, lista_dict_matricula_completa_alumnos, fuente_negrita, alineacion_centrada, borde_celda, relleno_encabezados)
            self.cargar_datos_dias_habiles(hoja, dias_habiles, fuente_negrita, alineacion_centrada, borde_celda, relleno_encabezados)
            self.cargar_encabezados_promedio_porcentaje_asistencia_inasistencia_alumnos(hoja, fuente_negrita, alineacion_centrada, borde_celda, relleno_encabezados)
            self.cargar_datos_promedio_asistencia_inasistencia_alumnos(hoja, lista_dict_promedio_asistencia_inasistencia_alumnos, alineacion_centrada, borde_celda)
            self.cargar_datos_porcentaje_asistencia_inasistencia_alumnos(hoja, lista_dict_porcentaje_asistencia_inasistencia_alumnos, alineacion_centrada, borde_celda)
            
            libro.save(ruta_archivo)
        except Exception as error:
            raise error


if __name__ == "__main__":
    """reporte_asistencia_mensual_alumnos = ReporteAsistenciaMensualAlumnos()
    
    asistencia_alumnos_repositorio = AsistenciaAlumnoRepositorio()
    especialidades_repositorio = EspecialidadRepositorio()
    
    asistencia_alumnos_servicio = AsistenciaAlumnoServicio(asistencia_alumnos_repositorio)
    especialidades_servicio = EspecialidadServicio(especialidades_repositorio)
    
    ESPECIALIDAD_ID = 1
    ANIO = 2024
    MES = 5
    
    datos = reporte_asistencia_mensual_alumnos.cargar_datos(asistencia_alumnos_repositorio, especialidades_repositorio, ESPECIALIDAD_ID, ANIO, MES)
    
    reporte_asistencia_mensual_alumnos.exportar(datos)"""
